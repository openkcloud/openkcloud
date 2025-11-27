"""
Excel Exporter - Export data to Excel format.

This module provides Excel export functionality with formatting.
Requires: openpyxl
"""

import io
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def export_to_excel(
    data: List[Dict[str, Any]],
    sheet_name: str = "Sheet1",
    columns: Optional[List[str]] = None,
    include_header: bool = True,
    auto_width: bool = True
) -> bytes:
    """
    Export data to Excel format.

    Args:
        data: List of dictionaries to export
        sheet_name: Worksheet name
        columns: Column names (if None, uses keys from first row)
        include_header: Whether to include header row
        auto_width: Auto-adjust column widths

    Returns:
        Excel file bytes

    Raises:
        ImportError: If openpyxl is not installed
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # Determine columns
    if columns is None:
        columns = list(data[0].keys())

    # Write header
    if include_header:
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            # Header formatting
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, row_data in enumerate(data, 2 if include_header else 1):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    if auto_width:
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2 if include_header else 1, len(data) + (2 if include_header else 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


def export_power_to_excel(power_data: Dict[str, Any]) -> bytes:
    """
    Export power monitoring data to Excel format with multiple sheets.

    Args:
        power_data: Power data dictionary

    Returns:
        Excel file bytes
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    timestamp = power_data.get('timestamp', datetime.utcnow().isoformat())

    # Summary sheet
    if 'summary' in power_data:
        ws_summary = wb.create_sheet("Summary")
        summary = power_data['summary']

        headers = ["Metric", "Value", "Unit"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        metrics = [
            ("Timestamp", timestamp, ""),
            ("Total Power", summary.get('total_power_watts', 0), "Watts"),
            ("Resource Count", summary.get('resource_count', 0), ""),
            ("Average Power", summary.get('avg_power_watts', 0), "Watts"),
            ("Max Power", summary.get('max_power_watts', 0), "Watts"),
            ("Min Power", summary.get('min_power_watts', 0), "Watts")
        ]

        for row_idx, (metric, value, unit) in enumerate(metrics, 2):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
            ws_summary.cell(row=row_idx, column=3, value=unit)

        # Auto-adjust widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 10

    # Breakdown sheet
    if 'breakdown' in power_data:
        breakdown_rows = []
        for item in power_data['breakdown']:
            breakdown_rows.append({
                'Name': item.get('name', ''),
                'Power (W)': item.get('power_watts', 0),
                'Percentage (%)': round(item.get('percentage', 0), 2),
                'Resource Count': item.get('resource_count', 0),
                'Resource Type': item.get('resource_type', '')
            })

        ws_breakdown = wb.create_sheet("Breakdown")
        columns = ['Name', 'Power (W)', 'Percentage (%)', 'Resource Count', 'Resource Type']

        # Header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws_breakdown.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Data
        for row_idx, row_data in enumerate(breakdown_rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws_breakdown.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))

        # Auto-adjust widths
        for col_idx in range(1, len(columns) + 1):
            ws_breakdown.column_dimensions[get_column_letter(col_idx)].width = 18

    # Accelerators sheet
    if 'accelerators' in power_data:
        ws_acc = wb.create_sheet("Accelerators")
        acc = power_data['accelerators']

        headers = ["Metric", "Value", "Unit"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_acc.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        metrics = [
            ("Total Power", acc.get('total_power_watts', 0), "Watts"),
            ("GPU Count", acc.get('gpu_count', 0), ""),
            ("NPU Count", acc.get('npu_count', 0), ""),
            ("Average GPU Power", acc.get('avg_gpu_power_watts', 0), "Watts")
        ]

        for row_idx, (metric, value, unit) in enumerate(metrics, 2):
            ws_acc.cell(row=row_idx, column=1, value=metric)
            ws_acc.cell(row=row_idx, column=2, value=value)
            ws_acc.cell(row=row_idx, column=3, value=unit)

        ws_acc.column_dimensions['A'].width = 20
        ws_acc.column_dimensions['B'].width = 15
        ws_acc.column_dimensions['C'].width = 10

    # Infrastructure sheet
    if 'infrastructure' in power_data:
        ws_infra = wb.create_sheet("Infrastructure")
        infra = power_data['infrastructure']

        headers = ["Metric", "Value", "Unit"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_infra.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        metrics = [
            ("Total Power", infra.get('total_power_watts', 0), "Watts"),
            ("Node Count", infra.get('node_count', 0), ""),
            ("Pod Count", infra.get('pod_count', 0), ""),
            ("Average Node Power", infra.get('avg_node_power_watts', 0), "Watts")
        ]

        for row_idx, (metric, value, unit) in enumerate(metrics, 2):
            ws_infra.cell(row=row_idx, column=1, value=metric)
            ws_infra.cell(row=row_idx, column=2, value=value)
            ws_infra.cell(row=row_idx, column=3, value=unit)

        ws_infra.column_dimensions['A'].width = 20
        ws_infra.column_dimensions['B'].width = 15
        ws_infra.column_dimensions['C'].width = 10

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


def export_metrics_to_excel(metrics_data: Dict[str, Any]) -> bytes:
    """
    Export performance metrics data to Excel format.

    Args:
        metrics_data: Metrics data dictionary

    Returns:
        Excel file bytes
    """
    rows = []
    timestamp = metrics_data.get('timestamp', datetime.utcnow().isoformat())

    # Handle different metric formats
    if 'metrics' in metrics_data:
        for metric in metrics_data['metrics']:
            rows.append({
                'Timestamp': timestamp,
                'Resource ID': metric.get('resource_id', ''),
                'Resource Type': metric.get('resource_type', ''),
                'Metric Name': metric.get('metric_name', ''),
                'Value': metric.get('value', 0),
                'Unit': metric.get('unit', ''),
                'Status': metric.get('status', '')
            })

    return export_to_excel(rows, sheet_name="Metrics")


def export_timeseries_to_excel(timeseries_data: Dict[str, Any]) -> bytes:
    """
    Export timeseries data to Excel format.

    Args:
        timeseries_data: Timeseries data dictionary

    Returns:
        Excel file bytes
    """
    rows = []

    if 'timeseries' in timeseries_data:
        metric_name = timeseries_data.get('metric_name', '')
        resource_type = timeseries_data.get('resource_type', '')

        for series in timeseries_data['timeseries']:
            resource_id = series.get('resource_id', series.get('name', ''))

            for datapoint in series.get('datapoints', []):
                timestamp, value = datapoint
                rows.append({
                    'Timestamp': timestamp,
                    'Resource ID': resource_id,
                    'Metric Name': metric_name,
                    'Value': value,
                    'Resource Type': resource_type
                })

    return export_to_excel(rows, sheet_name="Timeseries")
